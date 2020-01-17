#!/usr/bin/env python3
# coding: utf-8
#
import argparse
import re
import time
from collections import OrderedDict
import requests

from logzero import logger

import uiautomator2 as u2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                             PatternFill, Side)

DEBUG = True


def search_words(d: u2.Device, text: str):
    logger.debug("search: %s", text)
    d(resourceId="index-kw").clear_text()
    # d.xpath("取消").wait()
    d(resourceId="index-kw").click()
    d.send_keys(text)  # set_text(text)
    d.xpath("百度一下").wait()
    time.sleep(1.0)

    xpath = '//*[@resource-id="index-box"]//android.widget.Button'
    elements = d.xpath(xpath).all()
    if len(elements) < 5:
        time.sleep(2.0)
        elements = d.xpath(xpath).all()
        if elements:
            logger.info("Too faster for %s", text)
    if DEBUG:
        for el in elements:
            logger.debug("TEXT: %s", el.text)
    u2results = [el.text for el in elements]

    if not d.xpath("反馈").exists:
        d.swipe_ext("up", 0.5)
        elements = d.xpath(xpath).all()
        swipe_results = [el.text for el in elements]
        u2results = list(set(u2results + swipe_results))

    # if u2results != results:
    #     logger.info("1: %s", sorted(results))
    #     logger.info("2: %s", sorted(u2results))
        # raise RuntimeError()
    return u2results


def clean_text(name):
    return re.sub(r"\s", "", name).replace("/", "").replace("?", "")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("filename", help="input excel")
    args = parser.parse_args()

    d = u2.connect()  # ("192.168.0.28")
    d.set_fastinput_ime(True)
    d.set_orientation('natural')
    d.shell('am start -a android.intent.action.VIEW -d https://www.baidu.com')

    wb = load_workbook(filename=args.filename)
    sheet = wb[wb.sheetnames[0]]

    row = 1

    # skip useless lines
    for i in range(10):
        value = sheet.cell(row, 1).value
        if value == '产品名':
            row += 1
            break
        row += 1

    logger.info("read from excel")

    dicts = OrderedDict()
    product_names = dict()
    keyword = None
    product = None
    while True:
        _p = sheet.cell(row, 1).value
        if _p and _p.strip():
            product = clean_text(_p)

        _k = sheet.cell(row, 2).value
        if _k and _k.strip():
            keyword = clean_text(_k)
            product_names[keyword] = product

        prompts = []
        for col in [3, 4]:
            p = sheet.cell(row, col).value
            if p:
                if p == keyword:  # fix doc error
                    p = "DOC_ERROR_SAME"
                prompts.append(p)

        # when to the end of line
        if not _k and not prompts:
            break

        if keyword not in dicts:
            dicts[keyword] = []
        dicts[keyword].extend(prompts)
        row += 1

    # save result
    logger.info("save results to xlsx")
    wb = Workbook()
    sheet = wb.active
    row = 1
    count = 0
    for keyword, buy_prompts in dicts.items():
        # if row > 10:
        #     break
        count += 1
        #sheet.cell(row, 1).value = keyword
        logger.debug("search %s, %d/%d", keyword, count, len(dicts))
        baidu_prompts = list(search_words(d, keyword))
        product_name = product_names[keyword]

        exists = False
        start_row = row
        second_exists = False
        for text_idx, text in enumerate(buy_prompts):
            ok = False
            for bp in baidu_prompts:
                if clean_text(text) in clean_text(bp):
                    ok = True
            if ok and row - start_row < 2:
                if second_exists:
                    logger.warning("Skip second %s", text)
                    break
                if text_idx >= 1:
                    second_exists = True
                sheet.cell(row, 2).value = text
                row += 1
                exists = True

        if exists:
            sheet.cell(start_row, 1).value = keyword

            logger.info("%s: %d", keyword, row-start_row)
            # 合并单元格
            sheet.merge_cells(start_row=start_row,
                              start_column=1,
                              end_row=row-1,
                              end_column=1)
            top_left_cell = sheet.cell(start_row, 1)
            top_left_cell.alignment = Alignment(
                horizontal="center", vertical="center")

    wb.save(filename="saved.xlsx")


if __name__ == "__main__":
    main()
